import discord
from discord.ext import commands
from openpyxl import load_workbook
import datetime
import json
import random
import os

intents = discord.Intents.all()

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

bot = commands.Bot(command_prefix= '.', case_insensitive=True, intents = intents)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

@bot.event
async def on_ready():
    await bot.change_presence(status = discord.Status.online, activity=discord.Game('.help → get help'))
    print(">> Bot is online <<")
    
@bot.event
async def on_member_join(member: discord.Member):
    if member.guild.id == jdata["sever_ID"]:
        channel = bot.get_channel(jdata["welcome_channel"])
        await channel.send(f"**{member}**歡迎喵~")

@bot.event
async def on_member_remove(member: discord.Member):
    if member.guild.id == jdata["sever_ID"]:
        channel = bot.get_channel(jdata["welcome_channel"])
        await channel.send(f"抓到了喵!!**{member}**在偷尻喵~")

@bot.command()
@commands.is_owner()
async def load(ctx, extension):
    bot.load_extension(f'cmds.{extension}')
    await ctx.send(f'學習了 **{extension}** 指令包!!奇怪的知識增加了喵~')

@bot.command()
@commands.is_owner()
async def unload(ctx, extension):
    bot.unload_extension(f'cmds.{extension}')
    await ctx.send(f'少了 **{extension}** 指令包，本喵現在覺得一身輕喵~')

@bot.command()
@commands.is_owner()
async def reload(ctx, extension):
    bot.reload_extension(f'cmds.{extension}')
    await ctx.send(f'Re - Loaded {extension} done.')  

bot.remove_command('help')
@bot.command()
async def help(ctx, *comad):
    count(ctx)
    if comad:
        embed=discord.Embed(title="Command Query", color=0xffe26f, timestamp=datetime.datetime.utcnow())
        if comad[0] == 'arc':
            embed.add_field(name="Arcaea", value=".arc [difficulty...] \nRandom Arcaea songs\n抽取隨機Arcaea歌曲", inline=False)
        elif comad[0] == 'arcbomb':
            embed.add_field(name="Arcaea", value=".arcbomb <amount> [difficulty...] \nDraw 2~5 random Arcaea songs\n連續抽取2~5首Arcaea歌曲", inline=False)
        elif comad[0] == 'c2':
            embed.add_field(name="Cytus II", value=".c2 [difficulty...] \nRandom Cytus II songs\n抽取隨機Cytus II歌曲", inline=False)
        elif comad[0] == 'c2bomb':
            embed.add_field(name="Cytus II", value=".c2bomb <amount> [difficulty...] \nDraw 2~5 random Cytus II songs\n連續抽取2~5首Cytus II歌曲", inline=False)
        elif comad[0] == 'dee':
            embed.add_field(name="Deemo", value=".dee [difficulty...] \nRandom Deemo songs\n抽取隨機Deemo歌曲", inline=False)
        elif comad[0] == 'deebomb':
            embed.add_field(name="Deemo", value=".deebomb <amount> [difficulty...] \nDraw 2~5 random Deemo songs\n連續抽取2~5首Deemo歌曲", inline=False)
        elif comad[0] == 'member':
            embed.add_field(name="Common", value=".member \nQuery the number of servers (not including bots)\n顯示當前群組人數 (不包含機器人)", inline=False)
        elif comad[0] == 'mj':
            embed.add_field(name="Common", value=".mj <name> \nQuery members' joining time\n查詢成員進入伺服器時間", inline=False)
        elif comad[0] == 'ping':
            embed.add_field(name="Common", value=".ping \nQuery current bot delay\n查詢現在bot的延遲", inline=False)
        elif comad[0] == 'say':
            embed.add_field(name="Common", value=".say <msg> \nRepeat the entered text\n讓bot重複訊息", inline=False)
        elif comad[0] == 'choose':
            embed.add_field(name="Random", value=".choose [anything...] \nRandomly extract the entered text\n隨機選擇輸入的文字", inline=False)
        elif comad[0] == 'mora':
            embed.add_field(name="Random", value=".mora \nFinger-guessing game\n猜拳", inline=False)
        elif comad[0] == 'rs':
            embed.add_field(name="Random", value=".rs <num_of_people> <group_amount> <role> \nRandom team with a certain group of players\n隨機分隊", inline=False)
        elif comad[0] == 'rn':
            embed.add_field(name="Random", value=".rn <amount> <low> <top> \nRandom number extraction\n指定區間內隨機數字抽取", inline=False)
        elif comad[0] == 'cal':
            embed.add_field(name="Tool", value=".cal <formula> \nCalculate the answer\n計算算式的答案", inline=False)
        elif comad[0] == 'cnt':
            embed.add_field(name="Tool", value=".cnt <name> \nCheck the number of times you have used the commands\n查詢你到目前為止使用指令的次數", inline=False)
        elif comad[0] == 'roleinfo':
            embed.add_field(name="Tool", value=".roleinfo <role> \nGet the information of a role\n查詢身分組資訊", inline=False)
        elif comad[0] == 'color':
            embed.add_field(name="Tool", value=".color <color code> \nGet the concrete color of the input color code\n查詢輸入的色碼具體的顏色資訊", inline=False)
        elif comad[0] == 'pick':
            embed.add_field(name="Fun", value=".pick \n...\n使用這個指令展開你的礦工生活!", inline=False)
        elif comad[0] == 'sell':
            embed.add_field(name="Fun", value=".sell <mineral> <amount>\n...\n挖到的礦物可以變賣成<:Gcoin:736650411498406009>，擁有了<:Gcoin:736650411498406009>你可以去商店逛逛", inline=False)
        elif comad[0] == 'daily':
            embed.add_field(name="Fun", value=".daily \n...\n領取你的每日獎賞!!", inline=False)
        elif comad[0] == 'rob':
            embed.add_field(name="Fun", value=".rob <name>\n...\n向其他人搶奪財產!不過要注意，你可能會得面對喵警察的制裁", inline=False)
        elif comad[0] == 'bag':
            embed.add_field(name="Fun", value=".bag \n...\n查看你的背包(錢和寶物)", inline=False) 
        elif comad[0] == 'ibag':
            embed.add_field(name="Fun", value=".ibag \n...\n查看你的道具", inline=False) 
        elif comad[0] == 'shop':
            embed.add_field(name="Fun", value=".shop \n...\n開啟商店(黑市)列表", inline=False) 
        elif comad[0] == 'buy':
            embed.add_field(name="Fun", value=".buy <amount> <goods> \n...\n購買寶物或道具", inline=False) 
        elif comad[0] == 'give':
            embed.add_field(name="Fun", value=".give <name> <amount> <item> \n...\n贈送指定玩家錢或寶物", inline=False) 
        elif comad[0] == 'info':
            embed.add_field(name="Fun", value=".info <name(can be blank)>\n...\n查看玩家的基本訊息", inline=False) 
        else:
            embed.add_field(name="Not Found", value="No such command found\n查無此指令", inline=False)
    
        await ctx.send(embed=embed)    
        
    else:   
        embed=discord.Embed(title="桔喵Assistant", description="Command Query", color=0xffe26f, timestamp=datetime.datetime.utcnow())
        embed.set_thumbnail(url="https://i.imgur.com/pms8YGV.png")
        embed.add_field(name=":musical_note: Rhythm Games", value="arc | arcbomb | c2 | c2bomb | dee | deebomb", inline=False)
        embed.add_field(name=":star: Common", value="member | mj | ping | say", inline=False)
        embed.add_field(name=":grey_question: Random", value="choose | mora | rs | rn", inline=False)
        embed.add_field(name=":tools: Tool", value="cal | cnt | roleinfo | color", inline=False)
        embed.add_field(name="<:Gcoin:736650744861556749> Fun(BETA)", value="pick | sell | daily | rob | bag | ibag | shop | buy | give | info", inline=False)
        embed.set_footer(text="use .help <cmd> to query the command you want to know")
        await ctx.send(embed=embed)

@bot.event
async def on_command_error(ctx, error):
    embed=discord.Embed(title="ERROR", color=0xff0000)

    if isinstance(error, commands.errors.CommandNotFound):
        return
    elif isinstance(error, commands.errors.MissingRequiredArgument):
        embed.add_field(name="Missing Required Argument缺少必要參數", value="你什麼都不說是要本喵怎麼辦喵?", inline=False)
        embed.add_field(name="Error message", value=error, inline=False)
    elif isinstance(error, commands.errors.BadArgument):
        embed.add_field(name="Incorrect or Unknown Argument型態錯誤或不明的參數", value="找遍了本喵的資料庫也幫不了你了喵~", inline=False)
        embed.add_field(name="Error message", value=error, inline=False)
    elif isinstance(error, commands.errors.CommandInvokeError):
        embed.add_field(name="An Exception Occurred呼叫指令時出現例外狀況", value="你...你...你...罰你去看.help喵!!", inline=False)
        embed.add_field(name="Error message", value=error, inline=False)
    elif isinstance(error, commands.errors.CommandOnCooldown):
        embed.add_field(name="急什麼急喵?", value=error, inline=False)
    else:
        embed.add_field(name="ERROR", value="本喵什麼都不知道喵~", inline=False)
        embed.add_field(name="Error message", value=error, inline=False)

    embed.set_footer(text="use .help <cmd> to query the command you want to know")
    await ctx.send(embed=embed)

for filename in os.listdir('./cmds'):
    if filename.endswith('.py'):
        bot.load_extension(f'cmds.{filename[:-3]}') 

if __name__ == "__main__":
    bot.run(jdata["TOKEN"])