import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import datetime
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

class Tool(Cog_Extension):

    @commands.command()
    async def cal(self, ctx, formula):
        count(ctx)
        ans = eval(formula)
        embed=discord.Embed(title="The answer is", description = f':small_orange_diamond: **{ans}**',color=0xffe26f)
        await ctx.send(embed=embed)

    @commands.command()
    async def cnt(self, ctx, name: discord.Member):
        wb = load_workbook('cmdcount.xlsx')
        ws = wb.active
        if ws['A1'].value != None:
            a = int(ws['A1'].value)
            for i in range(a):
                if str(ws['B' + str(i+1)].value) == str(name.id):
                    if name.nick != None:
                        embed=discord.Embed(title="YO!!", description = f'{name.nick} have used commands **{ws["C" + str(i+1)].value}** times~',color=0xffe26f)
                    else:
                        embed=discord.Embed(title="YO!!", description = f'{name} have used commands **{ws["C" + str(i+1)].value}** times~',color=0xffe26f)
                    embed.set_footer(text="GOOD JOB MEOW!")
                    break
                else:
                    if i == (a-1):
                        if name.nick != None:
                            embed=discord.Embed(title="OH~", description = f'It seems that **{name.nick}** hasn\'t use commands yet meow!',color=0xffe26f)
                        else:
                            embed=discord.Embed(title="OH~", description = f'It seems that **{name}** hasn\'t use commands yet meow!',color=0xffe26f)
        else:
            embed=discord.Embed(title="Humm...", description = f'I can\'t find any user in my data meow',color=0xffe26f)

        await ctx.send(embed=embed)
        wb.close()

    @commands.command()
    async def roleinfo(self, ctx, role: discord.Role):
        count(ctx)
        rL = role.members
        if role.hoist:
            ish = 'yes'
        else:
            ish = 'no'
        if role.mentionable:
            imen = 'yes'
        else:
            imen = 'no'
        c = str(role.color)[1:]
        colo = 'https://www.color-hex.com/color/' + str(role.color)[1:]
        col = 'https://dummyimage.com/80x80/' + c.upper() + '/' + c.upper() + '.jpg'
        # datetime_str = role.created_at.strftime("%Y/%m/%d %H:%M:%S") 
        datetime_str = role.created_at.date()
        embed=discord.Embed(title=f'The information of {role.name}',color=role.color)
        embed.add_field(name=f'Name', value=role.name, inline=True)
        embed.add_field(name=f'ID', value=str(role.id), inline=True)
        embed.add_field(name="Color", value=f'[{str(role.color).upper()}]({colo})', inline=True)
        embed.set_thumbnail(url=col)
        embed.add_field(name=f'Mention', value=f'`{role.mention}`', inline=True)
        embed.add_field(name=f'Members', value=str(len(rL)), inline=True)
        embed.add_field(name=f'Postion', value=str(role.position), inline=True)
        embed.add_field(name=f'Hoisted', value=ish, inline=True)
        embed.add_field(name=f'Mentionable', value=imen, inline=True)
        embed.set_footer(text=f'Created At • {datetime_str}')
        await ctx.send(embed=embed)
    
    @commands.command()
    async def color(self, ctx, colo: discord.Color):
        count(ctx)
        c = str(colo)[1:]
        co = 'https://www.color-hex.com/color/' + str(colo)[1:]
        col = 'https://dummyimage.com/80x80/' + c.upper() + '/' + c.upper() + '.jpg'
        embed=discord.Embed(title=f'The concrete color of {colo}',color=colo)
        embed.add_field(name="Link", value=f'[{str(colo).upper()}]({co})', inline=True)
        embed.set_thumbnail(url=col)
        await ctx.send(embed=embed)



def setup(bot):
    bot.add_cog(Tool(bot))