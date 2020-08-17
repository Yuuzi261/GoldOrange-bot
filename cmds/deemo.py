import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

class Deemo(Cog_Extension):


    @commands.command()
    async def dee(self, ctx, *difficulty):
        count(ctx)
        allist = ["9", "10", "11", "ex"]
        diffs = []
        temp = []
        boo = False
        for x in difficulty:
            diffs.append(x)

        for x in diffs:
            if x not in allist:
                boo = True
                temp.append(x)

        diffs = list(set(diffs).difference(set(temp)))

        if diffs == [] and not("all" in difficulty or "ALL" in difficulty):
            embed=discord.Embed(title="ERROR", color=0xff0000)
            embed.add_field(name=".dee", value="Please enter a reasonable difficulty meow~ (9 ~ 11, you can also enter 'ex')\n請輸入合理的難度喵~(9 ~ 11,也可以輸入ex)", inline=False)
            embed.set_image(url=(jdata["giraffe"]))
            await ctx.send(embed=embed)
            return

        if "all" in difficulty or "ALL" in difficulty:
            resultdiff = random.choice(allist)
            boo = False
        else:
            resultdiff = random.choice(diffs)
        
        if boo:
            embed=discord.Embed(title="WARNING", color=0xff8000)
            embed.add_field(name=".dee", value="Amazing typing~ Unrecognizable input meow~\nKeep typing in a mess, it's none of my business meow!!\n出現了無法識別的難度喵~已自動排除了喵~\n你再亂打啊喵~本喵沒在理你的喵!!", inline=False)
            await ctx.send(embed=embed)
            
        keyword = "deemolv" + resultdiff + "songs"
        random_song = random.choice(jdata[keyword])

        embed=discord.Embed(title="Hard to make decision?選擇困難症喵?", description="Take my hand meow讓本喵來幫你喵:", color=0xffe26f)
        embed.set_image(url=(random_song))
        await ctx.send(embed=embed)

    @commands.command()
    async def deebomb(self, ctx, amount, *difficulty):
        count(ctx)
        try:
            amount = int(amount)
        except:
            embed=discord.Embed(title="ERROR", color=0xff0000)
            embed.add_field(name=".deebomb", value="Please enter a reasonable amount meow~(2 ~ 5)\n請輸入合理的數量喵~(2 ~ 5)", inline=False)
            embed.set_image(url=(jdata["giraffe"]))
            await ctx.send(embed=embed)
            return

        if 1 < amount <= 5:

            allist = ["9", "10", "11", "ex"]
            diffs = []
            temp = []
            boo = False
            for x in difficulty:
                diffs.append(x)

            for x in diffs:
                if x not in allist:
                    boo = True
                    temp.append(x)

            diffs = list(set(diffs).difference(set(temp)))

            if diffs == [] and not("all" in difficulty or "ALL" in difficulty):
                embed=discord.Embed(title="ERROR", color=0xff0000)
                embed.add_field(name=".deebomb", value="Please enter a reasonable difficulty meow~ (9 ~ 11, you can also enter 'ex')\n請輸入合理的難度喵~(9 ~ 11,也可以輸入ex)", inline=False)
                embed.set_image(url=(jdata["giraffe"]))
                await ctx.send(embed=embed)
                return

            # if ((amount == 4) or (amount == 5)) and (len(diffs) == 1) and (diffs[0] == "8"):
            #     await ctx.send("8級只有3首喵~\n")
            #     await ctx.send("<:Gcat3:711805083695710228>")
            #     return
            
            songs = []

            if "all" in difficulty or "ALL" in difficulty:
                boo = False
                nowdiff = random.choice(allist)
            else:
                nowdiff = random.choice(diffs)

            if boo:
                embed=discord.Embed(title="WARNING", color=0xff8000)
                embed.add_field(name=".deebomb", value="Amazing typing~ Unrecognizable input meow~\nKeep typing in a mess, it's none of my business meow!!\n出現了無法識別的難度喵~已自動排除了喵~\n你再亂打啊喵~本喵沒在理你的喵!!", inline=False)
                await ctx.send(embed=embed)

            nowkeyword = "deemolv" + nowdiff + "songs"
            now = random.choice(jdata[nowkeyword])
            songs.append(now)
            i = 0
            while i < (amount - 1):
                if "all" in difficulty or "ALL" in difficulty:
                    allist = ["9", "10", "11", "ex"]
                    nowdiff = random.choice(allist)
                else:
                    nowdiff = random.choice(diffs)
                nowkeyword = "deemolv" + nowdiff + "songs"
                now = random.choice(jdata[nowkeyword])
                if now in songs:
                    continue
                else:
                    songs.append(now)
                    i+=1

            i = 0
            for i in range(amount):
                if i == 0:
                    embed=discord.Embed(title="Hard to make decision?選擇困難症喵?", description="Take my hand meow讓本喵來幫你喵:", color=0xffe26f)
                else:
                    embed=discord.Embed(color=0xffe26f)
                random_song = songs[i]
                embed.set_image(url=(random_song + '.jpg'))
                await ctx.send(embed=embed)

        else:
            embed=discord.Embed(title="ERROR", color=0xff0000)
            embed.add_field(name=".deebomb", value="Please enter a reasonable amount meow~(2 ~ 5)\n請輸入合理的數量喵~(2 ~ 5)", inline=False)
            embed.set_image(url=(jdata["giraffe"]))
            await ctx.send(embed=embed)
            return


def setup(bot):
    bot.add_cog(Deemo(bot))