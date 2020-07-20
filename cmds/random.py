import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

class Random(Cog_Extension):

    @commands.command()
    async def mora(self, ctx):
        count(ctx)
        random_result = random.choice(jdata["mora"])
        embed=discord.Embed(title="本喵贏定了喵~", color=0xffe26f)
        embed.set_image(url=(random_result + '.jpg'))
        await ctx.send(embed=embed)

    @commands.command()
    async def choose(self, ctx, *songs):
        count(ctx)
        result = random.choice(songs)
        embed=discord.Embed(title="Result", description = f'I choice **{result}** meow~\n本喵決定選 **{result}** 喵~', color=0xffe26f)
        await ctx.send(embed=embed)

    @commands.command()
    async def rs(self, ctx, num_of_people, group_amount, role: discord.Role):
        count(ctx)
        num = int(num_of_people)
        gamount = int(group_amount)
        boo = True
        if num % gamount != 0:
            m = num%gamount
            boo = False
        people = []
        for member in ctx.guild.members:
            if role in member.roles:
                people.append(member.name)
        
        selected_people = random.sample(people, k=num)

        embed=discord.Embed(title="Random Team隨機分隊", color=0xffe26f)
        
        if boo:
            for squad in range(gamount):
                a = random.sample(selected_people, k = (num//gamount))
                msg = " "
                for name in a:
                    msg = msg + name + ' , '
                embed.add_field(name=f':small_orange_diamond: Team**{squad+1}**:', value=msg[:-3], inline=False)
                for name in a:
                    selected_people.remove(name)
        else:
            for squad in range(gamount):
                if squad != (gamount-1):
                    a = random.sample(selected_people, k = (num//gamount))
                else:
                    a = random.sample(selected_people, k = m)
                    
                msg = " "
                for name in a:
                    msg = msg + name + ' , '
                embed.add_field(name=f':small_orange_diamond: Team**{squad+1}**:', value=msg[:-3], inline=False)
                for name in a:
                    selected_people.remove(name)

        await ctx.send(embed=embed)

    @commands.command()
    async def rn(self, ctx, amount: int, low: int, top: int):
        count(ctx)
        embed=discord.Embed(title="Random Number Extraction隨機數字抽取", color=0xffe26f)
        for i in range(amount):
            r = random.randint(low, top)
            embed.add_field(name=f':small_orange_diamond: Result**{i+1}**:', value=r, inline=True)
        
        await ctx.send(embed=embed)



def setup(bot):
    bot.add_cog(Random(bot))