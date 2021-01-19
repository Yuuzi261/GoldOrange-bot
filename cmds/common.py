import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import time
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

class Common(Cog_Extension):

    @commands.command()
    async def ping(self, ctx):
        count(ctx)
        embed=discord.Embed(title=f'本喵現在的反應延遲是 **{round(self.bot.latency*1000)}** ms', description=f'至少比你的頭腦還要快上 **{round(250/(self.bot.latency*1000),3)}** 倍喵(・∀・)', color=0xffe26f)
        await ctx.send(embed=embed)

    @commands.command()
    async def spthx(self, ctx):
        count(ctx)
        embed=discord.Embed(title="【特別感謝】", description="協助開發者一覽", color=0xffe26f)
        embed.set_thumbnail(url="https://i.imgur.com/pms8YGV.png")
        embed.add_field(name="**本喵的主要開發者是 金桔 喵~**", value="目前沒有其它的共同開發者喵！\n特別感謝以下的人協助開發喵(ㆁωㆁ)", inline=False)
        embed.add_field(name="**:small_orange_diamond: 黑洞以及其他所有的大佬**", value="感謝協助Debug以及提供教學喵", inline=False)
        embed.add_field(name="**:small_orange_diamond: SJ**", value="感謝提供C2、Deemo所有歌曲圖片喵", inline=False)
        embed.add_field(name="**:small_orange_diamond: 小杰**", value="感謝提供arcaea所有歌曲圖片喵", inline=False)
        embed.add_field(name="**:small_orange_diamond: 所有參與測試人員**", value="感謝所有群內幫助測試以及提供意見的人員喵", inline=True)
        embed.set_footer(text="Programmer : 金桔")
        await ctx.send(embed=embed)

    @commands.command()
    async def say(self, ctx, *, msg):
        count(ctx)
        await ctx.message.delete()
        async with ctx.channel.typing():
            time.sleep(2)
            await ctx.send(msg)

    @commands.command() 
    async def member(self, ctx): 
        count(ctx)
        people = 0
        for member in ctx.guild.members: 
            if not(member.bot):
                people+=1

        embed=discord.Embed(title="People Counting伺服器人數統計",description = f'本群現在有 **{people}** 人喵~\nThere are **{people}** people in this server meow~', color=0xffe26f)
        embed.set_footer(text="The number of people has been automatically deducted from the number of bots meow~")
        await ctx.send(embed=embed)
        
    @commands.command() 
    async def mj(self, ctx, name: discord.Member = None): 
        count(ctx)
        if name == None:
            name = ctx.author
        for member in ctx.guild.members:
            if member == name:
                datetime_str = member.joined_at.date()
                embed=discord.Embed(title=f'{name.name}\'s joining time',description = f'{datetime_str}', color=0xffe26f)
                embed.set_footer(text="The displayed time is GMT standard time meow~")
                await ctx.send(embed=embed)


def setup(bot):
    bot.add_cog(Common(bot))