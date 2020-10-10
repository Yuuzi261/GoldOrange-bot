import discord
from discord.ext import commands
from openpyxl import load_workbook
from core.classes import Cog_Extension
import datetime
import random
import json

with open('setting.json', 'r', encoding='utf8') as jfile:
    jdata = json.load(jfile)

def count(c):
    wb = load_workbook('cmdcount.xlsx')
    ws = wb.active
    if ws['A1'].value != None:
        a = int(ws['A1'].value)
        for i in range(a):
            if str(ws['B' + str(i+1)].value) == str(c.author.id):
                ws['C' + str(i+1)].value = int(ws['C' + str(i+1)].value) + 1
                break
            else:
                if i == (a-1):
                    ws['A1'].value = int(ws['A1'].value) + 1
                    ws['B' + str(i+2)].value = str(c.author.id)
                    ws['C' + str(i+2)].value = 1
    else:
        ws['A1'].value = 1
        ws['B1'].value = str(c.author.id)
        ws['C1'].value = 1

    wb.save('cmdcount.xlsx')
    wb.close()

class Moderator(Cog_Extension):

    @commands.command()
    @commands.has_guild_permissions(administrator= True)
    async def roleclear(self, ctx, formula):
        #testing
        await ctx.send('有權限')


def setup(bot):
    bot.add_cog(Moderator(bot))